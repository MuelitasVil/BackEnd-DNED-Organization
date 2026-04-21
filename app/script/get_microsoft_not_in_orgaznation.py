# python -m app.script.get_microsoft_not_in_orgaznation --period 2026-1

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set

from openpyxl import Workbook, load_workbook
from sqlmodel import Session

from app.configuration.database import engine
from app.domain.models.user_unal import UserUnal
from app.service.crud.user_unal_service import UserUnalService
from app.utils.keyword_not_person import verify_is_person


EMAIL_COLUMN_INDEX = 3
DEFAULT_INPUT_NAME = "microsoft.xlsx"
EXTERNAL_USER_TAG = "#ext#"


@dataclass
class UserValidationResult:
    email: str
    in_db: bool
    is_person: Optional[bool]
    is_active_in_period: bool
    full_name: Optional[str]
    db_name: Optional[str]
    reason: str


def normalize_email(email: str) -> str:
    return str(email).strip().lower()


def get_input_rows(input_path: Path) -> List[tuple[str, Optional[str]]]:
    wb = load_workbook(filename=input_path, data_only=True)
    ws = wb.active

    rows: List[tuple[str, Optional[str]]] = []
    seen: Set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=2).value
        raw_email = ws.cell(row=row_idx, column=EMAIL_COLUMN_INDEX).value
        if not raw_email:
            continue

        email = normalize_email(raw_email)
        if EXTERNAL_USER_TAG in email:
            continue

        if not email or email in seen:
            continue

        seen.add(email)
        name = str(raw_name).strip() if raw_name else None
        rows.append((email, name))

    wb.close()
    return rows


def get_users_maps(
    session: Session,
    cod_period: str,
) -> tuple[Set[str], Dict[str, UserUnal]]:
    active_users = UserUnalService.get_all_by_period(cod_period, session)
    all_users = UserUnalService.get_all_no_pagination(session)

    active_set = {
        normalize_email(u.email_unal)
        for u in active_users
        if u.email_unal
    }
    all_users_map = {
        normalize_email(u.email_unal): u
        for u in all_users
        if u.email_unal
    }
    return active_set, all_users_map


def validate_users(
    input_rows: List[tuple[str, Optional[str]]],
    active_users: Set[str],
    all_users_map: Dict[str, UserUnal],
) -> List[UserValidationResult]:
    results: List[UserValidationResult] = []

    for email, excel_name in input_rows:
        db_user = all_users_map.get(email)
        in_db = db_user is not None
        db_name = get_display_name(db_user)
        full_name = excel_name or db_name
        is_person = verify_is_person(full_name) if full_name else False
        is_active_in_period = email in active_users

        reasons: List[str] = []
        if not is_active_in_period:
            reasons.append("NO_ACTIVO_EN_PERIODO")
        if not in_db:
            reasons.append("NO_REGISTRADO_EN_DB")
        if not is_person:
            reasons.append("NO_PERSONA")
        reason = ";".join(reasons) if reasons else "OK"

        result = UserValidationResult(
            email=email,
            in_db=in_db,
            is_person=is_person,
            is_active_in_period=is_active_in_period,
            full_name=full_name,
            db_name=db_name,
            reason=reason,
        )
        results.append(result)

    return results


def get_display_name(user: Optional[UserUnal]) -> Optional[str]:
    if not user:
        return None

    if user.full_name:
        return str(user.full_name).strip()

    name = str(user.name).strip() if user.name else ""
    lastname = str(user.lastname).strip() if user.lastname else ""
    joined = f"{name} {lastname}".strip()
    return joined or None


def write_output_excel(
    output_path: Path,
    cod_period: str,
    users: List[UserValidationResult],
) -> None:
    wb = Workbook()

    ws_result = wb.active
    ws_result.title = "resultado"
    ws_result.append([
        "email",
        "periodo",
        "nombre",
        "nombre en db",
        "esta_en_db",
        "activo_en_periodo",
        "es_persona",
        "motivo",
    ])

    for user in users:
        ws_result.append([
            user.email,
            cod_period,
            user.full_name,
            user.db_name,
            user.in_db,
            user.is_active_in_period,
            user.is_person,
            user.reason,
        ])

    wb.save(output_path)
    wb.close()


def build_default_output_path(base_dir: Path, cod_period: str) -> Path:
    safe_period = cod_period.replace("/", "-").replace(" ", "")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"microsoft_not_active_{safe_period}_{timestamp}.xlsx"
    return base_dir / file_name


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(
        description=(
            "Lee el archivo de Microsoft, valida usuarios no activos en un "
            "periodo y exporta un Excel de resultados."
        )
    )
    parser.add_argument(
        "--period",
        required=True,
        help="Codigo del periodo a validar. Ejemplo: 2025-2",
    )
    parser.add_argument(
        "--input",
        default=str(script_dir / DEFAULT_INPUT_NAME),
        help=(
            "Ruta del Excel de entrada. Por defecto: "
            "app/script/microsoft.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Ruta del Excel de salida. "
            "Si no se define, se crea automaticamente."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).resolve()
    if not input_path.exists():
        raise FileNotFoundError(
            f"No se encontro el archivo de entrada: {input_path}"
        )

    output_path = (
        Path(args.output).resolve()
        if args.output
        else build_default_output_path(input_path.parent, args.period)
    )

    with Session(engine) as session:
        input_rows = get_input_rows(input_path)
        active_users, all_users_map = get_users_maps(session, args.period)
        users = validate_users(
            input_rows,
            active_users,
            all_users_map,
        )

    write_output_excel(
        output_path=output_path,
        cod_period=args.period,
        users=users,
    )

    active_count = len([u for u in users if u.is_active_in_period])
    not_active_count = len([u for u in users if not u.is_active_in_period])
    not_person_count = len([u for u in users if not u.is_person])

    print(f"Total correos leidos: {len(input_rows)}")
    print(f"Usuarios activos en periodo {args.period}: {active_count}")
    print(f"Usuarios no activos en periodo {args.period}: {not_active_count}")
    print(f"Usuarios no persona detectados: {not_person_count}")
    print(f"Archivo generado: {output_path}")


if __name__ == "__main__":
    main()
